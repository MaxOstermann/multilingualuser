from django.contrib.auth import authenticate, login, logout
from rest_framework.decorators import api_view
from rest_framework.response import Response
from user.models import MultilingualUser
from user.serializers import MultilingualUserSerializer
from django.contrib.auth.decorators import login_required
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.translation import ugettext as _


@login_required
@api_view(['GET'])
def user_list_api(request):
    users = MultilingualUser.objects.all()
    if 'first_name' in request.data:
        users = users.filter(first_name=request.data['first_name'])
    if 'last_name' in request.data:
        users = users.filter(last_name=request.data['last_name'])
    if 'position' in request.data:
        users = users.filter(position=request.data['position'])
    if 'birthday_year' in request.data and 'birthday_month' in request.data:
        users = users.filter(
            birthday__year=request.data['birthday_year'],
            birthday__month=request.data['birthday_month']
        )
    serializer = MultilingualUserSerializer(users, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def login_api(request):
    user = authenticate(username=request.data['username'], password=request.data['password'])
    if user is not None:
        login(request, user)
        return Response("Вход выполнен.")
    else:
        return Response("Ваши логин или пароль не соответствуют.")


@api_view(['GET'])
def logout_api(request):
    logout(request)
    return Response("Выход выполнен.")


@api_view(['GET'])
def export_api(request):
    users_queryset = MultilingualUser.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Users'

    columns = [
        _('username'),
        _('first name'),
        _('last name'),
        _('position'),
        _('image id'),
        _('birthday'),
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for user in users_queryset:
        row_num += 1

        row = [
            user.username,
            user.first_name,
            user.last_name,
            user.position,
            user.image_id.url,
            str(user.birthday),
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
